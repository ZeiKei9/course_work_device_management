import csv
import json

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def export_service_orders_to_csv(queryset):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="service_orders.csv"'
    response.write("\ufeff".encode("utf8"))

    writer = csv.writer(response)
    writer.writerow(
        [
            "ID",
            "Device",
            "Issue Description",
            "Status",
            "Priority",
            "Assigned To",
            "Created By",
            "Created At",
        ]
    )

    for order in queryset:
        writer.writerow(
            [
                order.id,
                order.device.name,
                order.issue_description,
                order.get_status_display(),
                order.get_priority_display(),
                order.assigned_to.username if order.assigned_to else "",
                order.created_by.username if order.created_by else "",
                order.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            ]
        )

    return response


def export_service_orders_to_excel(queryset):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="service_orders.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Service Orders"

    headers = [
        "ID",
        "Device",
        "Issue Description",
        "Status",
        "Priority",
        "Assigned To",
        "Created By",
        "Created At",
    ]

    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_num, order in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1).value = order.id
        ws.cell(row=row_num, column=2).value = order.device.name
        ws.cell(row=row_num, column=3).value = order.issue_description
        ws.cell(row=row_num, column=4).value = order.get_status_display()
        ws.cell(row=row_num, column=5).value = order.get_priority_display()
        ws.cell(row=row_num, column=6).value = (
            order.assigned_to.username if order.assigned_to else ""
        )
        ws.cell(row=row_num, column=7).value = (
            order.created_by.username if order.created_by else ""
        )
        ws.cell(row=row_num, column=8).value = order.created_at.strftime(
            "%Y-%m-%d %H:%M:%S"
        )

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response


def export_service_orders_to_json(queryset):
    response = HttpResponse(content_type="application/json")
    response["Content-Disposition"] = 'attachment; filename="service_orders.json"'

    orders_data = []
    for order in queryset:
        order_dict = {
            "id": order.id,
            "device": {"id": order.device.id, "name": order.device.name},
            "issue_description": order.issue_description,
            "status": order.status,
            "status_display": order.get_status_display(),
            "priority": order.priority,
            "priority_display": order.get_priority_display(),
            "assigned_to": {
                "id": order.assigned_to.id,
                "username": order.assigned_to.username,
            }
            if order.assigned_to
            else None,
            "created_by": {
                "id": order.created_by.id,
                "username": order.created_by.username,
            }
            if order.created_by
            else None,
            "created_at": order.created_at.isoformat(),
        }
        orders_data.append(order_dict)

    response.write(json.dumps(orders_data, indent=2, ensure_ascii=False))
    return response

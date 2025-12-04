import csv
import json
import os

from django.conf import settings
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.fonts import addMapping

font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'arial.ttf')

pdfmetrics.registerFont(TTFont('RusFont', font_path))
addMapping('RusFont', 0, 0, 'RusFont')
addMapping('RusFont', 0, 1, 'RusFont')
addMapping('RusFont', 1, 0, 'RusFont')
addMapping('RusFont', 1, 1, 'RusFont')
font_name = 'RusFont'


# --- CSV EXPORT FUNCTIONS ---

def export_devices_to_csv(queryset):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="devices.csv"'
    response.write("\ufeff".encode("utf8"))

    writer = csv.writer(response)
    writer.writerow(
        [
            "ID",
            "Name",
            "Serial Number",
            "Inventory Number",
            "Category",
            "Brand",
            "Status",
            "Condition",
            "Location",
            "Purchase Date",
            "Purchase Price",
            "Warranty Until",
            "Created At",
        ]
    )

    for device in queryset:
        writer.writerow(
            [
                device.id,
                device.name,
                device.serial_number,
                device.inventory_number,
                device.category.name if device.category else "",
                device.brand.name if device.brand else "",
                device.get_status_display(),
                device.get_condition_display(),
                device.location.name if device.location else "",
                device.purchase_date,
                device.purchase_price,
                device.warranty_until,
                device.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            ]
        )

    return response


def export_loans_to_csv(queryset):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="loans.csv"'
    response.write("\ufeff".encode("utf8"))

    writer = csv.writer(response)
    writer.writerow(
        [
            "ID",
            "User",
            "Device",
            "Serial Number",
            "Manager",
            "Loaned At",
            "Due Date",
            "Status",
        ]
    )

    for loan in queryset:
        writer.writerow(
            [
                loan.id,
                loan.user.username,
                loan.device.name,
                loan.device.serial_number,
                loan.manager.username if loan.manager else "",
                loan.loaned_at.strftime("%Y-%m-%d %H:%M:%S"),
                loan.due_date.strftime("%Y-%m-%d %H:%M:%S"),
                loan.get_status_display(),
            ]
        )

    return response


def export_devices_to_excel(queryset):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="devices.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Devices"

    headers = [
        "ID",
        "Name",
        "Serial Number",
        "Inventory Number",
        "Category",
        "Brand",
        "Status",
        "Condition",
        "Location",
        "Purchase Date",
        "Purchase Price",
        "Warranty Until",
        "Created At",
    ]

    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_num, device in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1).value = device.id
        ws.cell(row=row_num, column=2).value = device.name
        ws.cell(row=row_num, column=3).value = device.serial_number
        ws.cell(row=row_num, column=4).value = device.inventory_number
        ws.cell(row=row_num, column=5).value = (
            device.category.name if device.category else ""
        )
        ws.cell(row=row_num, column=6).value = device.brand.name if device.brand else ""
        ws.cell(row=row_num, column=7).value = device.get_status_display()
        ws.cell(row=row_num, column=8).value = device.get_condition_display()
        ws.cell(row=row_num, column=9).value = (
            device.location.name if device.location else ""
        )
        ws.cell(row=row_num, column=10).value = device.purchase_date
        ws.cell(row=row_num, column=11).value = device.purchase_price
        ws.cell(row=row_num, column=12).value = device.warranty_until
        ws.cell(row=row_num, column=13).value = device.created_at.strftime(
            "%Y-%m-%d %H:%M:%S"
        )

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response


def export_loans_to_excel(queryset):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="loans.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Loans"

    headers = [
        "ID",
        "User",
        "Device",
        "Serial Number",
        "Manager",
        "Loaned At",
        "Due Date",
        "Status",
    ]

    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_num, loan in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1).value = loan.id
        ws.cell(row=row_num, column=2).value = loan.user.username
        ws.cell(row=row_num, column=3).value = loan.device.name
        ws.cell(row=row_num, column=4).value = loan.device.serial_number
        ws.cell(row=row_num, column=5).value = (
            loan.manager.username if loan.manager else ""
        )
        ws.cell(row=row_num, column=6).value = loan.loaned_at.strftime(
            "%Y-%m-%d %H:%M:%S"
        )
        ws.cell(row=row_num, column=7).value = loan.due_date.strftime(
            "%Y-%m-%d %H:%M:%S"
        )
        ws.cell(row=row_num, column=8).value = loan.get_status_display()

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response


# --- JSON EXPORT FUNCTIONS ---

def export_devices_to_json(queryset):
    response = HttpResponse(content_type="application/json")
    response["Content-Disposition"] = 'attachment; filename="devices.json"'

    devices_data = []
    for device in queryset:
        device_dict = {
            "id": device.id,
            "name": device.name,
            "serial_number": device.serial_number,
            "inventory_number": device.inventory_number,
            "category": {"id": device.category.id, "name": device.category.name}
            if device.category
            else None,
            "brand": {"id": device.brand.id, "name": device.brand.name}
            if device.brand
            else None,
            "status": device.status,
            "status_display": device.get_status_display(),
            "condition": device.condition,
            "condition_display": device.get_condition_display(),
            "location": {"id": device.location.id, "name": device.location.name}
            if device.location
            else None,
            "purchase_date": str(device.purchase_date)
            if device.purchase_date
            else None,
            "purchase_price": str(device.purchase_price)
            if device.purchase_price
            else None,
            "warranty_until": str(device.warranty_until)
            if device.warranty_until
            else None,
            "created_at": device.created_at.isoformat(),
        }
        devices_data.append(device_dict)

    response.write(json.dumps(devices_data, indent=2, ensure_ascii=False))
    return response


def export_loans_to_json(queryset):
    response = HttpResponse(content_type="application/json")
    response["Content-Disposition"] = 'attachment; filename="loans.json"'

    loans_data = []
    for loan in queryset:
        loan_dict = {
            "id": loan.id,
            "user": {
                "id": loan.user.id,
                "username": loan.user.username,
                "email": loan.user.email,
            },
            "device": {
                "id": loan.device.id,
                "name": loan.device.name,
                "serial_number": loan.device.serial_number,
            },
            "manager": {"id": loan.manager.id, "username": loan.manager.username}
            if loan.manager
            else None,
            "loaned_at": loan.loaned_at.isoformat(),
            "due_date": loan.due_date.isoformat(),
            "status": loan.status,
            "status_display": loan.get_status_display(),
            "notes": loan.notes,
        }
        loans_data.append(loan_dict)

    response.write(json.dumps(loans_data, indent=2, ensure_ascii=False))
    return response


# --- PDF EXPORT FUNCTIONS ---

def export_devices_to_pdf(queryset):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="devices.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    elements = []

    styles = getSampleStyleSheet()

    # Создаем свой стиль заголовка, чтобы не зависеть от Bold в Heading1
    header_style = ParagraphStyle(
        'RusHeader',
        parent=styles['Normal'],
        fontName=font_name, # Используем 'RusFont' или 'Helvetica'
        fontSize=14,
        alignment=1, # По центру
        spaceAfter=20
    )

    elements.append(Paragraph("Список устройств", header_style))

    # Данные
    data = [['ID', 'Название', 'Серийный №', 'Статус', 'Состояние']]
    for dev in queryset:
        data.append([
            str(dev.id),
            str(dev.name)[:25],
            str(dev.serial_number),
            dev.get_status_display(),
            dev.get_condition_display()
        ])

    table = Table(data, colWidths=[40, 200, 100, 100, 100])

    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name), # Используем наш шрифт
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))

    elements.append(table)
    doc.build(elements)
    return response

def export_loans_to_pdf(queryset):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="loans.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    elements = []

    styles = getSampleStyleSheet()

    header_style = ParagraphStyle(
        'RusHeader',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=14,
        alignment=1,
        spaceAfter=20
    )

    elements.append(Paragraph("Отчет по выдачам", header_style))

    data = [['ID', 'Сотрудник', 'Устройство', 'Дата выдачи', 'Дата возврата']]
    for loan in queryset:
        data.append([
            str(loan.id),
            loan.user.username,
            loan.device.name[:20],
            loan.loaned_at.strftime("%Y-%m-%d"),
            loan.due_date.strftime("%Y-%m-%d")
        ])

    table = Table(data, colWidths=[40, 150, 200, 100, 100])

    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    elements.append(table)
    doc.build(elements)
    return response
